from openpyxl import load_workbook, Workbook
# 1
# file_name = 'data.xlsx'
# workbook = load_workbook(file_name)
# list1 = []
# for sheetname in workbook.sheetnames:
#     worksheet = workbook[sheetname]
#     for j in range(1, worksheet.max_column+1):
#         val = worksheet.cell(1, j).value
#         if val is not None:
#             list1.append(val)
# print(list1)
# new_wb = Workbook()
# new_ws = new_wb.active
# for column_index, value in enumerate(list1, 1):
#     new_ws.cell(row=1, column=column_index, value=value)
# new_wb.save("new_data.xlsx")

#8
# x = input()
# if len(x) == 6:
#     first_part = sum(map(int, x[:3]))
#     second_part = sum(map(int, x[3:]))
#     if first_part == second_part:
#         print("Счастливый")
#     else:
#         print("Обычный")
# else:
#     print("Неверное количество знаков")

#9
a=int(input())
b=int(input())
s=1
while True:
    s+=1
    if s%a==0 and s%b==0:
        break
print(s)
